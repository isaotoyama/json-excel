import pandas as pd
import openpyxl
import json

def excel_to_json_with_reversed_hyperlinks(excel_file_path, json_file_path):
    try:
        # Load the workbook and get the sheet names
        wb = openpyxl.load_workbook(excel_file_path)
        sheet_names = wb.sheetnames
        
        json_data = []
        
        # Iterate over each sheet in the Excel file
        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            
            # Extract column headers
            columns = [cell.value for cell in ws[1]]
            
            # Initialize a list to hold the rows
            data = []
            
            # Iterate over the rows in the sheet
            for row in ws.iter_rows(min_row=2):
                row_data = {}
                for cell, column in zip(row, columns):
                    if column == 'Website Link' and cell.hyperlink:
                        row_data['filename'] = cell.value 
                        row_data['downloadLink'] = cell.hyperlink.target
                    else:
                        row_data[column.lower().replace(' ', '_')] = cell.value
                data.append(row_data)
            
            # Split the sheet name to get id and title
            id_title = sheet_name.split('_', 1)
            if len(id_title) < 2:
                continue
            
            item_id, title = id_title
            
            # Append to json_data
            json_data.append({
                'id': item_id,
                'title': title,
                'cards': data
            })
        
        # Write the JSON data to a file
        with open(json_file_path, 'w') as f:
            json.dump(json_data, f, indent=4)
        
        return f"Successfully converted {excel_file_path} to {json_file_path}"
    except Exception as e:
        return f"An error occurred: {e}"

# File paths for conversion
excel_file_path = 'files/zzz.xlsx'  # Path to the uploaded Excel file
json_file_path_reversed = 'files/zzz.json'  # Path to the desired JSON file

# Convert Excel file back to JSON with reversed hyperlinks
result_reversed = excel_to_json_with_reversed_hyperlinks(excel_file_path, json_file_path_reversed)
result_reversed
